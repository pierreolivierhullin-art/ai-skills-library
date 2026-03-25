#!/usr/bin/env python3
"""
generate_financial_template.py
-------------------------------
Génère un fichier Excel avec une structure de modèle financier standard.

Feuilles générées :
  - Hypothèses  : paramètres et drivers financiers
  - P&L         : compte de résultat sur 3 ans (N, N+1, N+2)
  - Bilan       : bilan comptable simplifié sur 3 ans
  - Cash Flow   : tableau des flux de trésorerie sur 3 ans
  - Dashboard   : synthèse des KPIs clés avec ratios

Dépendances :
  pip install openpyxl

Usage :
  python3 generate_financial_template.py [nom_fichier.xlsx]

Exemples :
  python3 generate_financial_template.py
  python3 generate_financial_template.py mon_modele_financier.xlsx
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Erreur : le module 'openpyxl' est requis.")
    print("Installation : pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constantes de style
# ---------------------------------------------------------------------------

COLOR_HEADER_DARK   = "FF1F3864"
COLOR_HEADER_BLUE   = "FF2E75B6"
COLOR_INPUT         = "FFD6E4F7"
COLOR_CALC          = "FFFFFFFF"
COLOR_OUTPUT        = "FFE2EFDA"
COLOR_TOTAL         = "FFBDD7EE"
COLOR_DASHBOARD_BG  = "FFF2F2F2"
COLOR_KPI_HEADER    = "FF375623"
COLOR_TEXT_WHITE    = "FFFFFFFF"
COLOR_TEXT_DARK     = "FF1F1F1F"

FONT_TITLE   = Font(name="Calibri", bold=True, size=14, color=COLOR_TEXT_WHITE)
FONT_SECTION = Font(name="Calibri", bold=True, size=11, color=COLOR_TEXT_WHITE)
FONT_HEADER  = Font(name="Calibri", bold=True, size=10, color=COLOR_TEXT_WHITE)
FONT_LABEL   = Font(name="Calibri", size=10, color=COLOR_TEXT_DARK)
FONT_FORMULA = Font(name="Calibri", size=10, italic=True, color="FF595959")
FONT_TOTAL   = Font(name="Calibri", bold=True, size=10, color=COLOR_TEXT_DARK)
FONT_NOTE    = Font(name="Calibri", size=8, italic=True, color="FF808080")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN  = Side(style="thin",   color="FFB0B0B0")
THICK = Side(style="medium", color="FF404040")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_EUR  = '#,##0 "€";[RED]-#,##0 "€"'
FMT_PCT  = '0.0%'
FMT_INT  = '#,##0'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def style_cell(cell, font=None, fill_color=None, alignment=None,
               border=None, number_format=None):
    if font:         cell.font = font
    if fill_color:   cell.fill = fill(fill_color)
    if alignment:    cell.alignment = alignment
    if border:       cell.border = border
    if number_format: cell.number_format = number_format


def header_row(ws, row, texts, bg, font, col_start=1):
    for i, txt in enumerate(texts):
        c = ws.cell(row=row, column=col_start + i, value=txt)
        style_cell(c, font=font, fill_color=bg, alignment=ALIGN_CENTER, border=BORDER_THIN)


def lbl(ws, row, col, text, indent=0, is_total=False, is_section=False):
    val = ("  " * indent + text) if indent else text
    c = ws.cell(row=row, column=col, value=val)
    if is_section:
        style_cell(c, font=FONT_SECTION, fill_color=COLOR_HEADER_BLUE,
                   alignment=ALIGN_LEFT, border=BORDER_THIN)
    elif is_total:
        style_cell(c, font=FONT_TOTAL, fill_color=COLOR_TOTAL,
                   alignment=ALIGN_LEFT, border=BORDER_THIN)
    else:
        style_cell(c, font=FONT_LABEL, fill_color=COLOR_INPUT,
                   alignment=ALIGN_LEFT, border=BORDER_THIN)


def val(ws, row, col, value=None, formula=None, fmt=FMT_EUR,
        is_total=False, bg=None):
    v = formula if formula else value
    c = ws.cell(row=row, column=col, value=v)
    fc = bg or (COLOR_TOTAL if is_total else COLOR_CALC)
    fn = FONT_TOTAL if is_total else (FONT_FORMULA if formula else FONT_LABEL)
    style_cell(c, font=fn, fill_color=fc, alignment=ALIGN_RIGHT,
               border=BORDER_THIN, number_format=fmt)


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col) if isinstance(col, int) else col].width = w


# ---------------------------------------------------------------------------
# Feuille : Hypothèses
# ---------------------------------------------------------------------------

def create_hypotheses(wb):
    ws = wb.create_sheet("Hypothèses")
    ws.tab_color = "2E75B6"
    set_widths(ws, {"A": 36, "B": 18, "C": 18, "D": 18, "E": 24})

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "MODÈLE FINANCIER — HYPOTHÈSES & PARAMÈTRES"
    style_cell(c, font=FONT_TITLE, fill_color=COLOR_HEADER_DARK, alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = "Cellules bleues = saisie. Cellules blanches = calculées automatiquement."
    style_cell(c, font=FONT_NOTE, alignment=ALIGN_CENTER)

    # Section Horizon
    row = 4
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = "HORIZON DE PROJECTION"
    style_cell(c, font=FONT_SECTION, fill_color=COLOR_HEADER_DARK,
               alignment=ALIGN_LEFT, border=BORDER_THIN)

    row += 1
    header_row(ws, row, ["Paramètre", "N", "N+1", "N+2", "Commentaire"],
               COLOR_HEADER_BLUE, FONT_HEADER)

    params = [
        ("Année de référence",       2024,  "=B7+1",  "=C7+1",  FMT_INT, "Modifier l'année de départ"),
        ("Taux de croissance CA",    0.15,  0.20,     0.18,     FMT_PCT, "Croissance annuelle du CA"),
        ("Taux de marge brute",      0.65,  0.67,     0.68,     FMT_PCT, "Marge brute = (CA - COGS) / CA"),
        ("Taux de charges fixes",    0.30,  0.28,     0.26,     FMT_PCT, "Charges fixes / CA"),
        ("Taux d'imposition IS",     0.25,  0.25,     0.25,     FMT_PCT, "Impôt sur les sociétés"),
    ]
    for i, (name, vn, vn1, vn2, fmt, note) in enumerate(params):
        r = row + 1 + i
        lbl(ws, r, 1, name)
        for col_idx, v_raw in [(2, vn), (3, vn1), (4, vn2)]:
            is_formula = isinstance(v_raw, str) and v_raw.startswith("=")
            if is_formula:
                val(ws, r, col_idx, formula=v_raw, fmt=fmt)
            else:
                val(ws, r, col_idx, value=v_raw, fmt=fmt, bg=COLOR_INPUT)
        note_cell = ws.cell(row=r, column=5, value=note)
        style_cell(note_cell, font=FONT_NOTE, alignment=ALIGN_LEFT)

    # Section CA
    row += len(params) + 2
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = "CHIFFRE D'AFFAIRES"
    style_cell(c, font=FONT_SECTION, fill_color=COLOR_HEADER_DARK,
               alignment=ALIGN_LEFT, border=BORDER_THIN)

    row += 1
    header_row(ws, row, ["Ligne de revenus", "N", "N+1", "N+2", "Commentaire"],
               COLOR_HEADER_BLUE, FONT_HEADER)

    ca_lines = [
        ("Produit / Service A", 500_000, "Principale source de revenus"),
        ("Produit / Service B", 200_000, "Deuxième ligne produit"),
        ("Produit / Service C", 100_000, "Troisième ligne produit"),
    ]
    first_ca = row + 1
    for i, (name, vn, note) in enumerate(ca_lines):
        r = first_ca + i
        lbl(ws, r, 1, name, indent=1)
        val(ws, r, 2, value=vn, fmt=FMT_EUR, bg=COLOR_INPUT)
        val(ws, r, 3, formula=f"=B{r}*(1+$C$9)", fmt=FMT_EUR)
        val(ws, r, 4, formula=f"=C{r}*(1+$D$9)", fmt=FMT_EUR)
        nc = ws.cell(row=r, column=5, value=note)
        style_cell(nc, font=FONT_NOTE, alignment=ALIGN_LEFT)

    total_ca = first_ca + len(ca_lines)
    lbl(ws, total_ca, 1, "TOTAL CHIFFRE D'AFFAIRES", is_total=True)
    for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
        val(ws, total_ca, ci,
            formula=f"=SUM({cl}{first_ca}:{cl}{total_ca-1})",
            fmt=FMT_EUR, is_total=True)

    ws.freeze_panes = "B6"
    ws.sheet_view.showGridLines = False
    return ws, total_ca


# ---------------------------------------------------------------------------
# Feuille : P&L
# ---------------------------------------------------------------------------

def create_pl(wb):
    ws = wb.create_sheet("P&L")
    ws.tab_color = "375623"
    set_widths(ws, {"A": 38, "B": 18, "C": 18, "D": 18})

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "COMPTE DE RÉSULTAT PRÉVISIONNEL"
    style_cell(c, font=FONT_TITLE, fill_color=COLOR_HEADER_DARK, alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 30

    row = 3
    header_row(ws, row, ["", "N", "N+1", "N+2"], COLOR_HEADER_BLUE, FONT_HEADER)

    # Structure P&L avec valeurs de départ
    structure = [
        # (label, indent, is_section, is_total, val_N)
        ("PRODUITS D'EXPLOITATION",              0, True,  False, None),
        ("Chiffre d'affaires net",               1, False, False, 800_000),
        ("Autres produits d'exploitation",       1, False, False, 5_000),
        ("TOTAL PRODUITS",                       0, False, True,  None),

        ("CHARGES D'EXPLOITATION",               0, True,  False, None),
        ("Coût des ventes (COGS)",               1, False, False, 280_000),
        ("Salaires et charges sociales",         1, False, False, 200_000),
        ("Frais marketing & acquisition",        1, False, False, 50_000),
        ("Dotations aux amortissements",         1, False, False, 20_000),
        ("Autres charges d'exploitation",        1, False, False, 45_000),
        ("TOTAL CHARGES",                        0, False, True,  None),

        ("RÉSULTAT D'EXPLOITATION (EBIT)",       0, False, True,  None),

        ("Produits financiers",                  1, False, False, 2_000),
        ("Charges financières",                  1, False, False, -5_000),
        ("RÉSULTAT AVANT IMPÔT (EBT)",           0, False, True,  None),

        ("Impôt sur les sociétés",               1, False, False, None),
        ("RÉSULTAT NET",                         0, False, True,  None),

        ("— Ratios —",                           0, True,  False, None),
        ("Taux de marge brute",                  1, False, False, None),
        ("Taux de marge EBIT",                   1, False, False, None),
        ("Taux de marge nette",                  1, False, False, None),
    ]

    row_map = {}
    current = row + 1
    input_rows = []
    total_rows = []

    for (label, indent, is_sec, is_tot, v_n) in structure:
        row_map[label] = current
        lbl(ws, current, 1, label, indent=indent, is_total=is_tot, is_section=is_sec)

        if is_sec:
            for ci in [2, 3, 4]:
                c = ws.cell(row=current, column=ci)
                style_cell(c, font=FONT_SECTION, fill_color=COLOR_HEADER_BLUE,
                           alignment=ALIGN_RIGHT, border=BORDER_THIN)
        elif is_tot:
            for ci in [2, 3, 4]:
                c = ws.cell(row=current, column=ci)
                style_cell(c, font=FONT_TOTAL, fill_color=COLOR_TOTAL,
                           alignment=ALIGN_RIGHT, border=BORDER_THIN,
                           number_format=FMT_EUR)
            total_rows.append(current)
        elif v_n is not None:
            val(ws, current, 2, value=v_n, fmt=FMT_EUR, bg=COLOR_INPUT)
            val(ws, current, 3, formula=f"=B{current}*1.10", fmt=FMT_EUR)
            val(ws, current, 4, formula=f"=C{current}*1.10", fmt=FMT_EUR)
            input_rows.append(current)
        else:
            for ci in [2, 3, 4]:
                c = ws.cell(row=current, column=ci)
                style_cell(c, font=FONT_FORMULA, fill_color=COLOR_CALC,
                           alignment=ALIGN_RIGHT, border=BORDER_THIN,
                           number_format=FMT_EUR)

        current += 1

    def fill_tot(label, formula_tpl):
        """formula_tpl uses _ as placeholder for column letter."""
        r = row_map.get(label)
        if not r:
            return
        for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
            ws.cell(row=r, column=ci).value = formula_tpl.replace("_", cl)

    r_ca   = row_map["Chiffre d'affaires net"]
    r_ap   = row_map["Autres produits d'exploitation"]
    r_tp   = row_map["TOTAL PRODUITS"]
    r_cogs = row_map["Coût des ventes (COGS)"]
    r_ac   = row_map["Autres charges d'exploitation"]
    r_tc   = row_map["TOTAL CHARGES"]
    r_ebit = row_map["RÉSULTAT D'EXPLOITATION (EBIT)"]
    r_pf   = row_map["Produits financiers"]
    r_cf   = row_map["Charges financières"]
    r_ebt  = row_map["RÉSULTAT AVANT IMPÔT (EBT)"]
    r_is   = row_map["Impôt sur les sociétés"]
    r_rn   = row_map["RÉSULTAT NET"]
    r_mgb  = row_map["Taux de marge brute"]
    r_ebitp= row_map["Taux de marge EBIT"]
    r_mnp  = row_map["Taux de marge nette"]

    fill_tot("TOTAL PRODUITS",
             f"=_{r_ca}+_{r_ap}")
    fill_tot("TOTAL CHARGES",
             f"=SUM(_{r_cogs}:_{r_ac})")
    fill_tot("RÉSULTAT D'EXPLOITATION (EBIT)",
             f"=_{r_tp}-_{r_tc}")
    fill_tot("RÉSULTAT AVANT IMPÔT (EBT)",
             f"=_{r_ebit}+_{r_pf}+_{r_cf}")

    for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
        # IS
        c = ws.cell(row=r_is, column=ci)
        c.value = f"=IF({cl}{r_ebt}>0,-{cl}{r_ebt}*0.25,0)"
        style_cell(c, font=FONT_FORMULA, fill_color=COLOR_CALC,
                   alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=FMT_EUR)
        # Résultat net
        c = ws.cell(row=r_rn, column=ci)
        c.value = f"={cl}{r_ebt}+{cl}{r_is}"
        style_cell(c, font=FONT_TOTAL, fill_color=COLOR_OUTPUT,
                   alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=FMT_EUR)
        # Ratios
        ca_ref = f"{cl}{r_ca}"
        c = ws.cell(row=r_mgb, column=ci)
        c.value = f"=IF({ca_ref}<>0,({ca_ref}-{cl}{r_cogs})/{ca_ref},0)"
        style_cell(c, font=FONT_FORMULA, fill_color=COLOR_CALC,
                   alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=FMT_PCT)
        c = ws.cell(row=r_ebitp, column=ci)
        c.value = f"=IF({ca_ref}<>0,{cl}{r_ebit}/{ca_ref},0)"
        style_cell(c, font=FONT_FORMULA, fill_color=COLOR_CALC,
                   alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=FMT_PCT)
        c = ws.cell(row=r_mnp, column=ci)
        c.value = f"=IF({ca_ref}<>0,{cl}{r_rn}/{ca_ref},0)"
        style_cell(c, font=FONT_FORMULA, fill_color=COLOR_CALC,
                   alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=FMT_PCT)

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    return ws, row_map


# ---------------------------------------------------------------------------
# Feuille : Bilan
# ---------------------------------------------------------------------------

def create_bilan(wb):
    ws = wb.create_sheet("Bilan")
    ws.tab_color = "ED7D31"
    set_widths(ws, {"A": 38, "B": 18, "C": 18, "D": 18})

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "BILAN PRÉVISIONNEL SIMPLIFIÉ"
    style_cell(c, font=FONT_TITLE, fill_color=COLOR_HEADER_DARK, alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 30

    row = 3
    header_row(ws, row, ["", "N", "N+1", "N+2"], COLOR_HEADER_BLUE, FONT_HEADER)

    def bilan_block(ws, start_row, items, samples):
        row_map = {}
        r = start_row
        for i, (label, is_sec, is_tot) in enumerate(items):
            row_map[label] = r
            lbl(ws, r, 1, label, is_total=is_tot, is_section=is_sec)
            v = samples[i]
            if v is not None and not is_sec and not is_tot:
                val(ws, r, 2, value=v, fmt=FMT_EUR, bg=COLOR_INPUT)
                val(ws, r, 3, formula=f"=B{r}*1.05", fmt=FMT_EUR)
                val(ws, r, 4, formula=f"=C{r}*1.05", fmt=FMT_EUR)
            elif is_sec:
                for ci in [2, 3, 4]:
                    style_cell(ws.cell(row=r, column=ci),
                               font=FONT_SECTION, fill_color=COLOR_HEADER_BLUE,
                               alignment=ALIGN_RIGHT, border=BORDER_THIN)
            elif is_tot:
                for ci in [2, 3, 4]:
                    style_cell(ws.cell(row=r, column=ci),
                               font=FONT_TOTAL, fill_color=COLOR_TOTAL,
                               alignment=ALIGN_RIGHT, border=BORDER_THIN,
                               number_format=FMT_EUR)
            r += 1
        return row_map, r

    actif_items = [
        ("ACTIF",                             True,  False),
        ("Actif immobilisé",                  True,  False),
        ("  Immobilisations corporelles (net)", False, False),
        ("  Immobilisations incorporelles",   False, False),
        ("  Immobilisations financières",     False, False),
        ("TOTAL ACTIF IMMOBILISÉ",            False, True),
        ("Actif circulant",                   True,  False),
        ("  Stocks",                          False, False),
        ("  Créances clients",                False, False),
        ("  Autres créances",                 False, False),
        ("  Trésorerie & équivalents",        False, False),
        ("TOTAL ACTIF CIRCULANT",             False, True),
        ("TOTAL ACTIF",                       False, True),
    ]
    actif_samples = [
        None, None, 150_000, 20_000, 10_000, None,
        None, 30_000, 80_000, 10_000, 50_000, None, None
    ]

    passif_items = [
        ("PASSIF",                            True,  False),
        ("Capitaux propres",                  True,  False),
        ("  Capital social",                  False, False),
        ("  Réserves",                        False, False),
        ("  Résultat de l'exercice",          False, False),
        ("TOTAL CAPITAUX PROPRES",            False, True),
        ("Dettes",                            True,  False),
        ("  Dettes financières (LT)",         False, False),
        ("  Dettes fournisseurs",             False, False),
        ("  Dettes fiscales et sociales",     False, False),
        ("  Autres dettes",                   False, False),
        ("TOTAL DETTES",                      False, True),
        ("TOTAL PASSIF",                      False, True),
    ]
    passif_samples = [
        None, None, 100_000, 50_000, 0, None,
        None, 80_000, 60_000, 30_000, 30_000, None, None
    ]

    actif_map, end_actif = bilan_block(ws, row + 1, actif_items, actif_samples)
    passif_map, end_passif = bilan_block(ws, end_actif + 1, passif_items, passif_samples)

    # Formules de totaux actif
    def sum_range(ws, total_row, first_data_row, last_data_row):
        for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
            ws.cell(row=total_row, column=ci).value = \
                f"=SUM({cl}{first_data_row}:{cl}{last_data_row})"

    r_immo_corps = actif_map.get("  Immobilisations corporelles (net)")
    r_immo_fin   = actif_map.get("  Immobilisations financières")
    r_tot_immo   = actif_map.get("TOTAL ACTIF IMMOBILISÉ")
    r_stocks     = actif_map.get("  Stocks")
    r_tresorerie = actif_map.get("  Trésorerie & équivalents")
    r_tot_circ   = actif_map.get("TOTAL ACTIF CIRCULANT")
    r_tot_actif  = actif_map.get("TOTAL ACTIF")

    if r_tot_immo:
        sum_range(ws, r_tot_immo, r_immo_corps, r_immo_fin)
    if r_tot_circ:
        sum_range(ws, r_tot_circ, r_stocks, r_tresorerie)
    if r_tot_actif and r_tot_immo and r_tot_circ:
        for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
            ws.cell(row=r_tot_actif, column=ci).value = \
                f"={cl}{r_tot_immo}+{cl}{r_tot_circ}"

    r_capital    = passif_map.get("  Capital social")
    r_resultat   = passif_map.get("  Résultat de l'exercice")
    r_tot_cp     = passif_map.get("TOTAL CAPITAUX PROPRES")
    r_dettes_lt  = passif_map.get("  Dettes financières (LT)")
    r_autres_dt  = passif_map.get("  Autres dettes")
    r_tot_dettes = passif_map.get("TOTAL DETTES")
    r_tot_passif = passif_map.get("TOTAL PASSIF")

    if r_tot_cp:
        sum_range(ws, r_tot_cp, r_capital, r_resultat)
    if r_tot_dettes:
        sum_range(ws, r_tot_dettes, r_dettes_lt, r_autres_dt)
    if r_tot_passif and r_tot_cp and r_tot_dettes:
        for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
            ws.cell(row=r_tot_passif, column=ci).value = \
                f"={cl}{r_tot_cp}+{cl}{r_tot_dettes}"

    # Vérification équilibre
    if r_tot_actif and r_tot_passif:
        check_row = end_passif + 1
        c = ws.cell(row=check_row, column=1, value="Contrôle Actif = Passif")
        style_cell(c, font=FONT_NOTE, alignment=ALIGN_LEFT)
        for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
            chk = ws.cell(row=check_row, column=ci)
            chk.value = (f'=IF({cl}{r_tot_actif}={cl}{r_tot_passif},"OK ✓","DÉSÉQUILIBRE !")')
            chk.font = Font(name="Calibri", bold=True, size=10, color="FF375623")
            chk.alignment = ALIGN_CENTER

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Feuille : Cash Flow
# ---------------------------------------------------------------------------

def create_cashflow(wb):
    ws = wb.create_sheet("Cash Flow")
    ws.tab_color = "7030A0"
    set_widths(ws, {"A": 42, "B": 18, "C": 18, "D": 18})

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "TABLEAU DES FLUX DE TRÉSORERIE (MÉTHODE INDIRECTE)"
    style_cell(c, font=FONT_TITLE, fill_color=COLOR_HEADER_DARK, alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 30

    row = 3
    header_row(ws, row, ["", "N", "N+1", "N+2"], COLOR_HEADER_BLUE, FONT_HEADER)

    cf_items = [
        # (label, is_sec, is_tot, val_N)
        ("I. FLUX D'EXPLOITATION (CFO)",              True,  False, None),
        ("Résultat net de l'exercice",                False, False, 80_000),
        ("+ Dotations aux amortissements",            False, False, 20_000),
        ("+ / - Variation du BFR",                   False, False, -10_000),
        ("FLUX DE TRÉSORERIE OPÉRATIONNELLE",         False, True,  None),

        ("II. FLUX D'INVESTISSEMENT (CFI)",           True,  False, None),
        ("Acquisitions d'immobilisations",            False, False, -50_000),
        ("Cessions d'actifs immobilisés",             False, False, 5_000),
        ("FLUX DE TRÉSORERIE D'INVESTISSEMENT",       False, True,  None),

        ("III. FLUX DE FINANCEMENT (CFF)",            True,  False, None),
        ("Augmentation de capital",                   False, False, 0),
        ("Nouveaux emprunts souscrits",               False, False, 30_000),
        ("Remboursements d'emprunts",                 False, False, -10_000),
        ("Dividendes distribués",                     False, False, 0),
        ("FLUX DE TRÉSORERIE DE FINANCEMENT",         False, True,  None),

        ("VARIATION NETTE DE TRÉSORERIE",             False, True,  None),
        ("Trésorerie d'ouverture",                    False, False, 20_000),
        ("TRÉSORERIE DE CLÔTURE",                     False, True,  None),
    ]

    row_map = {}
    current = row + 1
    cfo_data, cfi_data, cff_data = [], [], []
    section = None

    for (label, is_sec, is_tot, v_n) in cf_items:
        row_map[label] = current
        lbl(ws, current, 1, label, indent=0 if (is_sec or is_tot) else 1,
            is_total=is_tot, is_section=is_sec)
        if is_sec:
            for ci in [2, 3, 4]:
                style_cell(ws.cell(row=current, column=ci),
                           font=FONT_SECTION, fill_color=COLOR_HEADER_BLUE,
                           alignment=ALIGN_RIGHT, border=BORDER_THIN)
            if "I." in label:   section = "cfo"
            elif "II." in label: section = "cfi"
            elif "III." in label: section = "cff"
        elif is_tot:
            for ci in [2, 3, 4]:
                style_cell(ws.cell(row=current, column=ci),
                           font=FONT_TOTAL, fill_color=COLOR_TOTAL,
                           alignment=ALIGN_RIGHT, border=BORDER_THIN,
                           number_format=FMT_EUR)
        elif v_n is not None:
            val(ws, current, 2, value=v_n, fmt=FMT_EUR, bg=COLOR_INPUT)
            val(ws, current, 3, formula=f"=B{current}*1.10", fmt=FMT_EUR)
            val(ws, current, 4, formula=f"=C{current}*1.10", fmt=FMT_EUR)
            if section == "cfo":   cfo_data.append(current)
            elif section == "cfi": cfi_data.append(current)
            elif section == "cff": cff_data.append(current)
        current += 1

    def set_sum(label, data_rows):
        r = row_map.get(label)
        if not r: return
        for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
            refs = "+".join(f"{cl}{dr}" for dr in data_rows)
            ws.cell(row=r, column=ci).value = f"={refs}" if refs else "=0"

    set_sum("FLUX DE TRÉSORERIE OPÉRATIONNELLE",  cfo_data)
    set_sum("FLUX DE TRÉSORERIE D'INVESTISSEMENT", cfi_data)
    set_sum("FLUX DE TRÉSORERIE DE FINANCEMENT",   cff_data)

    r_cfo = row_map["FLUX DE TRÉSORERIE OPÉRATIONNELLE"]
    r_cfi = row_map["FLUX DE TRÉSORERIE D'INVESTISSEMENT"]
    r_cff = row_map["FLUX DE TRÉSORERIE DE FINANCEMENT"]
    r_var = row_map["VARIATION NETTE DE TRÉSORERIE"]
    r_ouv = row_map["Trésorerie d'ouverture"]
    r_clo = row_map["TRÉSORERIE DE CLÔTURE"]

    for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
        ws.cell(row=r_var, column=ci).value = \
            f"={cl}{r_cfo}+{cl}{r_cfi}+{cl}{r_cff}"
        prev_cl = get_column_letter(ci - 1)
        if ci == 2:
            ws.cell(row=r_clo, column=ci).value = f"={cl}{r_ouv}+{cl}{r_var}"
        else:
            ws.cell(row=r_clo, column=ci).value = f"={prev_cl}{r_clo}+{cl}{r_var}"
        style_cell(ws.cell(row=r_clo, column=ci),
                   font=FONT_TOTAL, fill_color=COLOR_OUTPUT,
                   alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=FMT_EUR)

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Feuille : Dashboard
# ---------------------------------------------------------------------------

def create_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.tab_color = "C00000"
    set_widths(ws, {"A": 3, "B": 28, "C": 16, "D": 16, "E": 16,
                     "F": 3, "G": 28, "H": 16, "I": 16, "J": 16, "K": 3})

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "DASHBOARD FINANCIER — KPIs CLÉS"
    style_cell(c, font=FONT_TITLE, fill_color=COLOR_HEADER_DARK, alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = ("Dashboard alimenté par les feuilles P&L, Bilan et Cash Flow. "
               "Ne pas modifier les cellules calculées.")
    style_cell(c, font=FONT_NOTE, alignment=ALIGN_CENTER)

    # KPI blocks : (titre, col_start, couleur_titre, liste de KPIs)
    blocks = [
        {
            "title": "RENTABILITÉ & CROISSANCE",
            "col":   2,
            "color": COLOR_KPI_HEADER,
            "kpis": [
                ("CA N",                   "='P&L'!B5",  FMT_EUR),
                ("CA N+1",                 "='P&L'!C5",  FMT_EUR),
                ("CA N+2",                 "='P&L'!D5",  FMT_EUR),
                ("Croissance CA N→N+1",    "=IF('P&L'!B5<>0,('P&L'!C5-'P&L'!B5)/'P&L'!B5,0)", FMT_PCT),
                ("Marge brute N",          "='P&L'!B19", FMT_PCT),
                ("Marge EBIT N",           "='P&L'!B20", FMT_PCT),
                ("Marge nette N",          "='P&L'!B21", FMT_PCT),
                ("Résultat net N",         "='P&L'!B17", FMT_EUR),
            ]
        },
        {
            "title": "TRÉSORERIE",
            "col":   7,
            "color": "FF7030A0",
            "kpis": [
                ("Trésorerie clôture N",    "='Cash Flow'!B17", FMT_EUR),
                ("Flux opérationnel N",     "='Cash Flow'!B5",  FMT_EUR),
                ("Flux investissement N",   "='Cash Flow'!B9",  FMT_EUR),
                ("Flux financement N",      "='Cash Flow'!B15", FMT_EUR),
                ("Variation trésorerie N",  "='Cash Flow'!B16", FMT_EUR),
            ]
        },
    ]

    ROW_START = 4
    for block in blocks:
        cs = block["col"]
        row = ROW_START

        ws.merge_cells(start_row=row, start_column=cs,
                       end_row=row, end_column=cs + 1)
        c = ws.cell(row=row, column=cs, value=block["title"])
        style_cell(c, font=FONT_SECTION, fill_color=block["color"],
                   alignment=ALIGN_CENTER, border=BORDER_THIN)

        row += 1
        for j, hdr in enumerate(["KPI", "Valeur"]):
            c = ws.cell(row=row, column=cs + j, value=hdr)
            style_cell(c, font=FONT_HEADER, fill_color=COLOR_HEADER_BLUE,
                       alignment=ALIGN_CENTER, border=BORDER_THIN)

        for k, (kpi_name, formula, fmt) in enumerate(block["kpis"]):
            r = row + 1 + k
            c = ws.cell(row=r, column=cs, value=kpi_name)
            style_cell(c, font=FONT_LABEL, fill_color=COLOR_DASHBOARD_BG,
                       alignment=ALIGN_LEFT, border=BORDER_THIN)
            c = ws.cell(row=r, column=cs + 1, value=formula)
            style_cell(c, font=FONT_FORMULA, fill_color=COLOR_OUTPUT,
                       alignment=ALIGN_RIGHT, border=BORDER_THIN, number_format=fmt)

    # Note
    note_row = ROW_START + 12
    ws.merge_cells(f"A{note_row}:K{note_row}")
    c = ws[f"A{note_row}"]
    c.value = ("Note : Les références de cellules P&L / Cash Flow correspondent aux "
               "lignes générées. Ajuster si la structure du modèle est modifiée.")
    style_cell(c, font=FONT_NOTE, alignment=ALIGN_LEFT)

    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    output = sys.argv[1] if len(sys.argv) > 1 else "modele_financier.xlsx"
    if not output.endswith(".xlsx"):
        output += ".xlsx"

    print(f"Generation du modele financier : {output}")

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("  • Hypotheses...")
    create_hypotheses(wb)
    print("  • P&L...")
    create_pl(wb)
    print("  • Bilan...")
    create_bilan(wb)
    print("  • Cash Flow...")
    create_cashflow(wb)
    print("  • Dashboard...")
    create_dashboard(wb)

    wb.active = wb["Dashboard"]

    try:
        wb.save(output)
        p = Path(output).resolve()
        print(f"\n[OK] Fichier genere : {p}")
        print("\nFeuilles :")
        for name in wb.sheetnames:
            print(f"  - {name}")
        print("\nProchaines etapes :")
        print("  1. Ouvrir dans Excel / LibreOffice Calc")
        print("  2. Saisir les hypotheses (cellules bleues) dans la feuille Hypotheses")
        print("  3. Ajuster les valeurs N dans P&L, Bilan et Cash Flow")
        print("  4. Verifier l'equilibre du Bilan (ligne Controle)")
        print("  5. Consulter le Dashboard pour les KPIs synthetiques")
    except PermissionError:
        print(f"\n[ERREUR] Impossible d'ecrire '{output}'. "
              "Verifier que le fichier n'est pas ouvert.")
        sys.exit(1)


if __name__ == "__main__":
    main()
